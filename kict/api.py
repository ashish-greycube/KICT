import frappe
from frappe import _
import erpnext
from kict.kict.doctype.railway_receipt.railway_receipt import get_available_batches
from erpnext.stock.get_item_details import get_item_details,get_basic_details,get_price_list_rate_for
from frappe.model.mapper import get_mapped_doc
from frappe.utils import getdate,add_days,get_time,get_first_day,get_last_day,cint,flt,cstr,get_datetime,getdate
from kict.kict.report.royalty_storage.royalty_storage import get_item_price
from kict.kict.doctype.vessel.vessel import get_unique_item
from kict.kict.report.royalty_storage.royalty_storage import get_royalty_storage_items_and_rate
from erpnext.controllers.accounts_controller import get_taxes_and_charges
from frappe.utils import get_link_to_form
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
import os

def set_cargo_handling_option_name_and_is_periodic(self,method):
    for row in self.get("custom_cargo_handling_charges_slots_details"):
        option = str(row.percent_billing)+ "%" + " - " +row.billing_when
        row.cargo_handling_option_name = option
        billing_type = row.billing_when
        posotion_of_string_for_periodic = billing_type.find("Periodic")
        posotion_of_string_for_dispatch = billing_type.find("Dispatch")

        if posotion_of_string_for_periodic >= 0:
            row.billing_type = "Periodic"
        elif posotion_of_string_for_dispatch >= 0:
            row.billing_type = "Dispatch"
        else:
            row.billing_type = "Non-Periodic"

def validate_rate_percent_billing(self,method):
    total_percent = 0
    cargo_customer_group = frappe.db.get_single_value("Coal Settings","customer_group_for_cargo_customer")
    if self.customer_group == cargo_customer_group:
        if len(self.get("custom_cargo_handling_charges_slots_details")) == 0:
            frappe.throw(_("Please fill cargo handling charges slots details"))
        else:
            for row in self.get("custom_cargo_handling_charges_slots_details"):
                total_percent = total_percent + row.percent_billing
        
            if total_percent != 100:
                frappe.throw(_("Total of rate percent billing must be 100%"))
            else:
                pass
            
            if self.custom_storage_charge_based_on == "":
                frappe.throw(_("Please select storage charge based on type"))
            else:
                if self.custom_storage_charge_based_on == "Fixed Days":
                    if self.custom_no_of_days == None:
                        frappe.throw(_("Please add fixed number of days"))
                if self.custom_storage_charge_based_on == "Actual Storage Days":
                    if self.custom_free_storage_days == None:
                        frappe.throw(_("Please add number of free storage days"))
                    if len(self.get("custom_chargeable_storage_charges_slots_details")) == 0:
                        frappe.throw(_("Please fill chargeable storage charges slots details"))
                    else:
                        chargeable_slots_details = self.get("custom_chargeable_storage_charges_slots_details")
                        for row in chargeable_slots_details:
                            if row.item == None:
                                frappe.throw(_("Row {0} : Please select item").format(row.idx))
                    if len(self.get("custom_chargeable_storage_charges_slots_details")) < 2:
                        frappe.throw(_("Please add row in chargeable storage charges slots details"))


def change_status_for_dn_creation_in_railway_receipt_on_cancel_of_dn(self,method):
    frappe.db.set_value("Railway Receipt Item Details",self.custom_railway_receipt_detail,"is_dn_created","No")

def change_status_for_is_billed_in_on_cancel_of_si(self,method):
    if self.custom_type_of_cargo_handling_invoice and self.custom_type_of_cargo_handling_invoice=="Periodic":    
        participating_rr = frappe.db.get_all("Railway Receipt Item Details",
                                            filters = {"sales_invoice_reference":self.name},
                                            fields = ["name","is_billed","sales_invoice_reference"])
        for ele in participating_rr:
            frappe.db.set_value("Railway Receipt Item Details",ele.name,"is_billed","No")
            frappe.db.set_value("Railway Receipt Item Details",ele.name,"sales_invoice_reference","")

def copy_vessel_to_stock_entry_item(self,method):
    vessel=self.custom_vessel
    if vessel:
        for item in self.get("items"):
            item.to_vessel=vessel

def copy_source_vessel_to_stock_entry_item(self,method):
    vessel=self.custom_vessel
    if vessel:
        for item in self.get("items"):
            item.vessel=vessel

def set_batch_no_and_warehouse_for_handling_loss_audit_sortage(self,method):
    stock_entry_type_for_audit_shortage = frappe.db.get_single_value("Coal Settings","audit_shortage")
    stock_entry_type_for_handling_loss = frappe.db.get_single_value("Coal Settings","handling_loss")
    if self.stock_entry_type==stock_entry_type_for_audit_shortage or self.stock_entry_type==stock_entry_type_for_handling_loss:
        copy_source_vessel_to_stock_entry_item(self,method)
        item_table_with_batches=[]
        item_rows_to_be_removed=[]
        for item in self.get("items"):
            if item.batch_no!=None:
                pass
            elif item.batch_no==None:    
                args=frappe._dict({'item_code': item.item_code, 
                'has_serial_no': '0', 
                'has_batch_no': '1', 
                'qty':item.qty, 
                'based_on': 'FIFO', 
                'vessel':self.custom_vessel,
                # 'posting_date': self.posting_date,
                # 'posting_time': self.posting_time,
                'cmd': "kict.kict.doctype.railway_receipt.railway_receipt.get_auto_data"
                }) 
                available_batches=get_available_batches(args)               
                # check if batch total qty is less than required qty
                float_precision = cint(frappe.db.get_default("float_precision")) or 3
                qty_from_batches=0
                for batch in available_batches:
                    qty_from_batches=flt((qty_from_batches+batch.qty),float_precision)
                if len(available_batches) < 1:
                    table_html = frappe.bold("No batches found.")
                    msg="The qty available from batches is {0}, whereas required qty is {1}. Hence cannot proceed.".format(frappe.bold(qty_from_batches),frappe.bold(item.qty))	
                    frappe.throw(_(table_html+"<br>"+msg))				
                if item.qty>qty_from_batches:
                    table_body="<table border='1'><tr><td><b>Batch</b></td><td><b>Warehouse</b></td><td><b>Qty</b></td></tr>"
                    table_row=""
                    for item in available_batches:
                        table_row=table_row+"<tr><td>"+item.batch_no+"</td><td>"+item.warehouse+"</td><td>"+cstr(item.qty)+"</td></tr>"
                    table_html=table_body+table_row+"</table>"			
                    msg="The qty available from batches is {0}, whereas required qty is {1}. Hence cannot proceed.".format(frappe.bold(qty_from_batches),frappe.bold(item.qty))	
                    frappe.throw(_(table_html+"<br>"+msg))
                # for each batch, warehouse add seperate line item
                for batch in available_batches:
                    item_with_batch={'item_code':item.item_code,
                                     'item_name':item.item_name,
                                     'custom_customer':item.custom_customer,
                                     'description':item.description,
                                     'item_group':item.item_group,
                                     'uom':item.uom,
                                     'stock_uom':item.stock_uom,
                                     'transfer_qty':batch.qty,
                                     'conversion_factor':item.conversion_factor,
                                     'qty':batch.qty,
                                     'batch_no':batch.batch_no,
                                     's_warehouse':batch.warehouse,
                                     'vessel':self.custom_vessel,
                                     'use_serial_batch_fields':1}
                    item_table_with_batches.append(item_with_batch)
                
                item_rows_to_be_removed.append(item)
               
        if len(item_table_with_batches)>0:
            [self.items.remove(d) for d in item_rows_to_be_removed]
            for row in item_table_with_batches:
                frappe.msgprint(_('Item {0} has now batch {1} and source warehouse {2}').format(row.get('item_code'),row.get('batch_no'),row.get('s_warehouse')),alert=True)
                self.append("items",row)

def generate_and_set_batch_no(self,method):
    if self.stock_entry_type=="Cargo Received":
        copy_vessel_to_stock_entry_item(self,method)
        for item in self.get("items"):
            if item.batch_no==None:
                batch_name=check_batch_no_exist(item.to_vessel,item.item_code,self.posting_date,self.posting_time)
                if batch_name!=None:
                    item.use_serial_batch_fields=1
                    item.batch_no=batch_name
                    frappe.msgprint(_('Item {0} existing  batch {1} is added').format(item.item_code,item.batch_no),alert=True)
                else:
                    batch_name=generate_batch_no(item.to_vessel,item.item_code,self.posting_date,self.posting_time)
                    item.use_serial_batch_fields=1
                    item.batch_no=batch_name
                    frappe.msgprint(_('Item {0} new batch {1} is added').format(item.item_code,item.batch_no),alert=True)

def check_batch_no_exist(vessel,item_code,posting_date,posting_time):
    port_date=get_port_date(posting_date,posting_time)
    batch_name=None
    batch_found = frappe.db.get_list("Batch", 
            filters={"item":item_code,"custom_vessel":vessel,"manufacturing_date":port_date},
            fields=["name"]) 
    if len(batch_found)>0:
        batch_name=batch_found[0].name
    return batch_name

    

def generate_batch_no(vessel,item_code,posting_date,posting_time):
    port_date=get_port_date(posting_date,posting_time)
    is_customer_provided_item = frappe.db.get_value('Item', item_code, 'is_customer_provided_item')
    if is_customer_provided_item==0:
        return
    def make_batch(kwargs):
        if frappe.db.get_value("Item", kwargs.item, "has_batch_no"):
            if frappe.db.get_value("Item", kwargs.item, "is_customer_provided_item"):
                if frappe.db.get_value("Item", kwargs.item, "create_new_batch")==0:
                    kwargs.doctype = "Batch"
                    return frappe.get_doc(kwargs).insert().name
        
    batch_id=get_name_from_hash(port_date,item_code)
    batch_name= make_batch(frappe._dict({
                "batch_id":batch_id,
                "item": item_code,
                "custom_vessel":vessel,
                "manufacturing_date":port_date,
                "stock_uom":frappe.db.get_value('Item', item_code, 'stock_uom'),
                "use_batchwise_valuation":0
            }
        )
    )    
    return batch_name

def get_name_from_hash(port_date,item_code):
    current_date=getdate(port_date).strftime("%d.%m.%y")
    print('current_date',current_date)
    coal_commodity = frappe.db.get_value("Item",item_code,"custom_coal_commodity")
    if not coal_commodity:
        frappe.throw(_("Item {0} does not have coal commodity. It is required to generate batch no.").format(frappe.bold(item_code)))
    commodity_grade = frappe.db.get_value("Item",item_code,"custom_commodity_grade")
    if not commodity_grade:
        frappe.throw(_("Item {0} does not have commodity grade. It is required to generate batch no.").format(frappe.bold(item_code)))
    customer_abbreviation = frappe.db.get_value("Item",item_code,"custom_customer_abbreviation")
    if not customer_abbreviation:
        frappe.throw(_("Item {0} does not have customer abbreviation. It is required to generate batch no.").format(frappe.bold(item_code)))	
    """
    Get a name for a Batch by generating a unique hash.
    :return: The hash that was generated.
    """
    temp = None
    while not temp:
        temp = coal_commodity[:1].upper()+commodity_grade[:1].upper()+customer_abbreviation+"-"+current_date+"-"+frappe.generate_hash()[:2].upper()
        if frappe.db.exists("Batch", temp):
            temp = None
    return temp

def set_grt_billed_for_bh_in_vessel_detail_on_submit_of_si(self,method):
    if self.custom_type_of_invoice == "Berth Hire Charges":
        vessel_doc = frappe.get_doc("Vessel",self.vessel)
        customer_specific_total_tonnage = calculate_customer_specific_total_tonnage(self,method)
        for row in vessel_doc.get("vessel_details"):
            if row.customer_name==self.customer:
                if not self.custom_quantity_in_mt:
                    frappe.throw(_("Quantity in MT value in Sales Invoice is 0. Kindly put value in it"))
                billed_qty = (flt((customer_specific_total_tonnage),3) / flt((self.custom_quantity_in_mt),3) ) * flt((row.tonnage_mt),3)
                row.grt_billed_for_bh_for_si = billed_qty

        vessel_doc.save(ignore_permissions = True)

def set_grt_billed_for_bh_in_vessel_detail_on_submit_of_pi(self,method):
    if self.custom_type_of_invoice == "Berth Hire Charges":
        vessel_doc = frappe.get_doc("Vessel",self.vessel)
        customer_specific_total_tonnage = calculate_customer_specific_total_tonnage(self,method)
        for row in vessel_doc.get("vessel_details"):
            if row.customer_name==self.customer:
                billed_qty = (flt((customer_specific_total_tonnage),3) / flt((self.custom_quantity_in_mt),3) ) * flt((row.tonnage_mt),3)
                row.grt_billed_for_bh_for_pi = billed_qty

        vessel_doc.save(ignore_permissions = True)

def calculate_customer_specific_total_tonnage(self,method):
    vessel_doc = frappe.get_doc("Vessel",self.vessel)
    customer_specific_total_tonnage = 0
    for item in vessel_doc.get("vessel_details"):
        if item.customer_name==self.customer:
            customer_specific_total_tonnage = customer_specific_total_tonnage + item.tonnage_mt
    return customer_specific_total_tonnage


def validate_items_for_cargoreceived_handlingloss_auditsortage(self,method):
    cargo_received,handling_loss,audit_shortage = frappe.db.get_value("Coal Settings","Coal Settings",["cargo_received","handling_loss","audit_shortage"])
    stock_entry_type = self.get("stock_entry_type")
    if stock_entry_type ==cargo_received or stock_entry_type ==handling_loss or stock_entry_type ==audit_shortage:	
        for item in self.items:
            is_customer_provided_item = frappe.db.get_value('Item', item.item_code, 'is_customer_provided_item')
            if is_customer_provided_item == 0:
                frappe.throw(_("Row {0} has item {1}. It should be customer provided item as stock entry type is {2}.").format(frappe.bold(item.idx),item.item_code,stock_entry_type))			

def validate_vessel_is_not_closed_in_stock_entry(self,method):
    vessel_name = self.custom_vessel
    vessel_closure_value = frappe.db.get_value("Vessel",vessel_name,"vessel_closure")
    if vessel_closure_value == 1:
        frappe.throw(_("Vessel {0} is closed. You can not create Stock Entry.").format(frappe.bold(vessel_name)))

    for item in self.items:
        item_detail = frappe.db.get_all("Vessel Details",
                                        parent_doctype = "Vessel",
                                        filters={"parent":vessel_name,"item":item.item_code},
                                        fields=["cargo_closure"])
        if len(item_detail)>0:
            if item_detail[0].cargo_closure == 1:
                frappe.throw(_("Row# {0} : vessel {1} and cargo item {2} is closed. You can not create stock entry.")
                             .format(item.idx,frappe.bold(vessel_name),frappe.bold(item.item_code)))

def validate_vessel_is_not_closed_in_delivery_note(self,method):
    for item in self.items:
        vessel_name = item.vessel
        vessel_closure_value = frappe.db.get_value("Vessel",vessel_name,"vessel_closure")
        if vessel_closure_value == 1:
            frappe.throw(_("Vessel {0} is closed. You can not create Stock Entry.").format(frappe.bold(vessel_name)))        
        vessel_item = frappe.db.get_all("Vessel Details",
                                        parent_doctype = "Vessel",
                                        filters={"parent":vessel_name,"item":item.item_code},
                                        fields=["cargo_closure"])
        if len(vessel_item)>0:
            if vessel_item[0].cargo_closure == 1:
                frappe.throw(_("Row #{0}: vessel {1} and item {2} is closed. You can not create Delivery Note.")
                             .format(item.idx,frappe.bold(vessel_name),frappe.bold(item.item_code)))
                
def remove_calculation_for_percent_billing_on_cancel_of_si(self,method):
    if self.custom_type_of_invoice == "Berth Hire Charges":
        vessel_doc = frappe.get_doc("Vessel",self.vessel)
        for row in vessel_doc.get("vessel_details"):
            if row.customer_name==self.customer:
                row.grt_billed_for_bh_for_si = 0

        vessel_doc.save(ignore_permissions = True)

def remove_calculation_for_percent_billing_on_cancel_of_pi(self,method):
    if self.custom_type_of_invoice == "Berth Hire Charges":
        vessel_doc = frappe.get_doc("Vessel",self.vessel)
        for row in vessel_doc.get("vessel_details"):
            if row.customer_name==self.customer:
                row.grt_billed_for_bh_for_pi = 0

        vessel_doc.save(ignore_permissions = True)

def get_port_date(date,time):
    time=get_time(time)
    if time <= get_time('06:00:00'):
        return add_days(date,-1)
    elif time > get_time('06:00:00') and time <= get_time('23:59:59'):
        return date
    
@frappe.whitelist()
def create_purchase_invoice_for_royalty_charges(source_name=None,target_doc=None,supplier_name=None,supplier_invoice_no=None,posting_date=None):
    frappe.enqueue( method='kict.api._create_purchase_invoice_for_royalty_charges',
        queue="long",
        timeout=2000,
		is_async=True,
        job_name="create_purchase_invoice_for_royalty_charges",
        enqueue_after_commit= True,
        **{"source_name":source_name,
                 "target_doc":target_doc,
                 "supplier_name":supplier_name,
                 "supplier_invoice_no":supplier_invoice_no,
                 "posting_date":posting_date}
        )
    frappe.msgprint(_('Purchase Invoice for Royalty Charges is being created. Please check the comment section after 30 mins.'),
        alert=True,
        indicator='green')
    

@frappe.whitelist()
def _create_purchase_invoice_for_royalty_charges(source_name=None,target_doc=None,supplier_name=None,supplier_invoice_no=None,posting_date=None):
    try:
        comment_1=[]
        comment_2=[]
        comment_3=[]
        comment_4=[]
        comment_5=[]
        comment_6=[]
        source_name = frappe.get_last_doc('Vessel').name
        first_slot_item,first_slot_storage_charges,second_slot_item,second_slot_storage_charges = get_royalty_storage_items_and_rate(type="PI",transaction_date=posting_date)
        distinct_vessel_list,free_vessel_list,charged_vessel = get_data_from_royalty_charges_report_for_vessel(posting_date,type="Comment")
    
        def set_missing_values(source, target):
            
            eligible_vessels = []
            royalty_charges_vessel_billed=[]
            royalty_charges_report_data = get_data_from_royalty_charges_report_for_vessel(posting_date,type="PI")
            target.supplier = supplier_name
            target.bill_no = supplier_invoice_no
            target.custom_is_royalty_invoice = 1
            price_list= frappe.db.get_single_value("Coal Settings", "royalty_price_list")
            currency= frappe.db.get_value("Supplier", {"name": supplier_name}, ["default_currency"])
            if price_list:
                target.buying_price_list = price_list
                pass
            if currency:
                target.currency = currency

            target.set_posting_time = 1
            target.posting_date = posting_date

            filter__from_date=get_first_day(posting_date)
            filter__to_date=get_last_day(posting_date)
            posting_date_month = getdate(posting_date).month
            filter_date=add_days(posting_date,-32)
            sof_list = frappe.db.get_all("Statement of Fact",
                                        or_filters={'first_line_ashore':['between',[filter__from_date,filter__to_date]],
                                                'all_line_cast_off':['between',[filter__from_date,filter__to_date]] },          
                                    fields=["name","first_line_ashore","all_line_cast_off","current_month_stay_hours"])
            # print('sof_list',sof_list)  
            for sof in sof_list:
                vessel_closure = frappe.db.get_value('Vessel', sof.name, 'vessel_closure')
                if vessel_closure==0:
                    if sof.first_line_ashore and sof.all_line_cast_off:
                        first_line_ashore_month,all_line_cast_off_month = (sof.first_line_ashore).month,(sof.all_line_cast_off).month
                        if posting_date_month==first_line_ashore_month or posting_date_month==all_line_cast_off_month:
                            eligible_vessels.append(sof.name)
            # print('sof_list2',eligible_vessels)  
            if len(eligible_vessels)==0:
                frappe.throw("No eligible vessels found for {0} month".format(posting_date_month))
            # self.bill_no = ''
            # print("Condition satisfy",eligible_vessels)
            # alert 1
            comment_1.append("[1/5] : Eligible vessels identified are <b>{0}<b>".format(len(eligible_vessels)))
            # frappe.msgprint(_("[1/5] : Eligible vessels identified are <b>{0}<b>").format(len(eligible_vessels)),alert=True)        

            for vessel in eligible_vessels:
                # print(vessel)
                vessel_doc = frappe.get_doc("Vessel",vessel)
                type_of_vessel = vessel_doc.costal_foreign_vessle
                if type_of_vessel == "Foreign":
                    royalty_invoice_item = frappe.db.get_single_value("Coal Settings","birth_hire_item_for_foreign_vessel")
                if type_of_vessel == "Coastal":
                    royalty_invoice_item = frappe.db.get_single_value("Coal Settings","birth_hire_item_for_coastal_vessel")
                
                vessel_item_commodity = frappe.db.get_all("Vessel Details",
                                                        parent_doctype = "Vessel",
                                                        filters={"parent":vessel},
                                                        fields=["coal_commodity"])
                distinct_item_commodity = []
                if len(vessel_item_commodity)>0:
                    for row in vessel_item_commodity:
                        if row.coal_commodity not in distinct_item_commodity:
                            distinct_item_commodity.append(row.coal_commodity)

                # print(distinct_item_commodity,"distinct_item_commodity")
                vessel_item_list = ",".join((ele if ele!=None else '') for ele in distinct_item_commodity)
                comment_2.append("[2/5] : Berth Hire : Vessel items charged are <b>{0}<b> ".format(vessel_item_list))
                # frappe.msgprint(_("[2/5] : Berth Hire : Vessel items charged are <b>{0}<b> ").format(vessel_item_list),alert=True)
                # royalty charges for berth hire
                bh_rate = get_item_price_list_rate(vessel,royalty_invoice_item,price_list,posting_date)

                royalty_percentage = frappe.db.get_single_value("Coal Settings","royalty_percentage")
                # custom_amount = custom_qty * bh_rate
                calculated_rate = bh_rate * (royalty_percentage/100)
                sof = frappe.db.get_all("Statement of Fact",
                            filters={"name":vessel},
                            fields=["vessel_stay_hours","current_month_stay_hours","next_month_stay_hours","first_line_ashore","all_line_cast_off"])
                actual_berth_hours = None
                current_month_stay_hours = 0
                vessel_next_month_stay_hours_is_zero=False
                # print(sof)
                if len(sof)>0:
                    first_line_month = (sof[0].first_line_ashore).month
                    all_line_month = (sof[0].all_line_cast_off).month
                    if posting_date_month==first_line_month and posting_date_month==all_line_month:
                        current_month_stay_hours = sof[0].current_month_stay_hours
                        actual_berth_hours = sof[0].vessel_stay_hours
                    elif posting_date_month == first_line_month:
                        current_month_stay_hours = sof[0].current_month_stay_hours
                        actual_berth_hours = sof[0].current_month_stay_hours
                    elif posting_date_month == all_line_month:
                        if sof[0].next_month_stay_hours > 0:
                            current_month_stay_hours = sof[0].next_month_stay_hours
                            actual_berth_hours = sof[0].next_month_stay_hours
                        else:
                            vessel_next_month_stay_hours_is_zero=True
                
                if vessel_next_month_stay_hours_is_zero==False:
                    custom_qty = vessel_doc.grt * current_month_stay_hours

                    item_details= {"item_code":royalty_invoice_item,"custom_vessel_name":vessel_doc.vessel_name,"custom_commodity":vessel_item_list,"custom_grt":vessel_doc.grt,"qty":custom_qty,
                    "custom_current_month_stay_hours":current_month_stay_hours,"custom_custom_qty":custom_qty,"rate":calculated_rate,"custom_actual_berth_hours":actual_berth_hours,"vessel":vessel}
                    purchase_invoice_item_bh = target.append("items",item_details)
                    if custom_qty==0:
                        frappe.throw(_("<b>Item quantity cannot be zero</b> <br> {0}".format(item_details)))
                    print("="*10,"berth hire")
                    print({"item_code":royalty_invoice_item,"custom_vessel_name":vessel_doc.vessel_name,"custom_commodity":vessel_item_list,"custom_grt":vessel_doc.grt,"qty":custom_qty,
                    "custom_current_month_stay_hours":current_month_stay_hours,"custom_custom_qty":custom_qty,"rate":calculated_rate,"custom_actual_berth_hours":actual_berth_hours,"vessel":vessel})
                    comment_3.append("[3/5] : Berth Hire : Qty is <b>{0}<b> and Rate is <b>{1}</b>".format(custom_qty,calculated_rate))
                    # frappe.msgprint(_("[3/5] : Berth Hire : Qty is <b>{0}<b> and Rate is <b>{1}</b>").format(custom_qty,calculated_rate),alert=True)
                # royalty charges for cargo handling
                royalty_invoice_item = frappe.db.get_single_value("Coal Settings","ch_charges")
                cargo_handling_data = get_cargo_handling_qty_for_royalty_purchase_invoice(posting_date)
                if len(cargo_handling_data)>0:
                    comment_4.append("[4/5] : Cargo Handling : Items idenfied for charging are <b>{0}<b>".format(len(cargo_handling_data)))
                    # frappe.msgprint(_("[4/5] : Cargo Handling : Items idenfied for charging are <b>{0}<b>").format(len(cargo_handling_data)),alert=True)
                    print("="*10,"cargo handling")
                    for row in cargo_handling_data:
                        if row.get("vessel") == vessel:
                            ch_price_list_rate = get_item_price_list_rate(vessel,royalty_invoice_item,price_list,posting_date)
                            qty = vessel_doc.total_tonnage_mt
                            # ch_amount = qty * ch_price_list_rate
                            ch_rate = ch_price_list_rate * (royalty_percentage/100)

                            purchase_invoice_item_ch =target.append("items",
                            {"item_code":royalty_invoice_item,"custom_vessel_name":vessel_doc.vessel_name,"custom_commodity":row.get("item"),"custom_grt":vessel_doc.total_tonnage_mt,"qty":row.get("qty"),
                            "custom_current_month_stay_hours":current_month_stay_hours,"custom_custom_qty":vessel_doc.total_tonnage_mt,"rate":ch_rate,"custom_actual_berth_hours":actual_berth_hours,"vessel":vessel})
                            print({"item_code":royalty_invoice_item,"custom_vessel_name":vessel_doc.vessel_name,"custom_commodity":row.get("item"),"custom_grt":vessel_doc.total_tonnage_mt,"qty":row.get("qty"),
                            "custom_current_month_stay_hours":current_month_stay_hours,"custom_custom_qty":vessel_doc.total_tonnage_mt,"rate":ch_rate,"custom_actual_berth_hours":actual_berth_hours,"vessel":vessel})
                # royalty charges for storage
                
                if len(royalty_charges_report_data)>0:
                    print("="*10,"storage")
                    comment_5.append("[5/5] : Storage : Items idenfied for charging are <b>{0}<b>".format(len(royalty_charges_report_data)))
                    # frappe.msgprint(_("[5/5] : Storage : Items idenfied for charging are <b>{0}<b>").format(len(royalty_charges_report_data)),alert=True)
                    
                    for row in royalty_charges_report_data:
                        if row.get("vessel") == vessel:
                            royalty_charges_vessel_billed.append(vessel)
                            if  row.get("storage_item") == first_slot_item:
                                item_commodity = frappe.db.get_value("Item",row.get("customer_item"),"custom_coal_commodity")
                                purchase_invoice_item_sc = target.append("items",
                                {"item_code":first_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":first_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":vessel,"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})
                                print( {"item_code":first_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":first_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":vessel,"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})
                            elif row.get("storage_item") == second_slot_item:
                                item_commodity = frappe.db.get_value("Item",row.get("customer_item"),"custom_coal_commodity")
                                purchase_invoice_item_sc_2 = target.append("items",
                                {"item_code":second_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":second_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":vessel,"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})
                                print({"item_code":second_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":second_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":vessel,"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})

            # for storage vessel which are not part of outer eligible vessels
            print('royalty_charges_report_data,royalty_charges_vessel_billed')
            print('1',royalty_charges_report_data,'2',royalty_charges_vessel_billed)
            if royalty_charges_report_data and len(royalty_charges_report_data)>0 and royalty_charges_vessel_billed and len(royalty_charges_vessel_billed)>0:
                print('#'*100)
                for row in royalty_charges_report_data:       
                    if row.get("vessel") not in royalty_charges_vessel_billed:
                        vessel_closure = frappe.db.get_value('Vessel', row.get("vessel"), 'vessel_closure')
                        if vessel_closure==0:
                            print("="*10,"storage old")
                            comment_6.append("[6] : Storage : Very old items that are not part of eligible vessels are <b>{0}<b>".format(row.get("vessel")))
                            # frappe.msgprint(_("[6] : Storage : Very old items that are not part of eligible vessels are <b>{0}<b>").format(row.get("vessel")),indicator="orange",alert=True)
                            if  row.get("storage_item") == first_slot_item:
                                item_commodity = frappe.db.get_value("Item",row.get("customer_item"),"custom_coal_commodity")
                                purchase_invoice_item_sc = target.append("items",
                                {"item_code":first_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":first_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":row.get("vessel"),"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})
                                print({"item_code":first_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":first_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":row.get("vessel"),"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})
                            elif row.get("storage_item") == second_slot_item:
                                item_commodity = frappe.db.get_value("Item",row.get("customer_item"),"custom_coal_commodity")
                                purchase_invoice_item_sc_2 = target.append("items",
                                {"item_code":second_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":second_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":row.get("vessel"),"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})     
                                print({"item_code":second_slot_item,"custom_vessel_name":row.get("vessel"),"custom_grt":vessel_doc.total_tonnage_mt,"rate":second_slot_storage_charges*(royalty_percentage/100),
                                "qty":row.get("qty"),"vessel":row.get("vessel"),"custom_for_stock_item":row.get("customer_item"),"custom_commodity":item_commodity})               

            default_taxes_template = frappe.db.get_single_value("Coal Settings","default_purchase_taxes_and_charges_template")
            if default_taxes_template:
                taxes = get_taxes_and_charges("Purchase Taxes and Charges Template", default_taxes_template)
                for tax in taxes:
                    target.append("taxes", tax)
            else : 
                frappe.throw(_("Please set Default Purchase Taxes and Charges Template in Coal Settings"))

        doc = get_mapped_doc('Vessel', source_name, {
            'Vessel': {
                'doctype': 'Purchase Invoice',
                # 'field_map': {
                # 	'agent_name':'name',
                # },			
                'validation': {
                    'docstatus': ['!=', 2]
                }
            },
            "Vessel Details": {
                "doctype": "Purchase Invoice Item",
                "condition":lambda doc:len(doc.name)<0,
                # "postprocess":update_item
            },	
        }, target_doc,set_missing_values)
        doc.run_method("set_missing_values")
        doc.run_method("calculate_taxes_and_totals")
        for ite in doc.get("items"):
            print(ite.get("name",ite.get("vessel"),ite.get("rate"),ite.get("qty")))
        doc.save()
        doc.add_comment("Comment", "<b>Eligible vessels</b> : {0} <br><hr><b>Free vessels</b> : {1} <br><hr><b>Charged vessels</b> : {2}".format(distinct_vessel_list,free_vessel_list,charged_vessel))
        print("comment_1,comment_2,comment_3,comment_4,comment_5,comment_6")
        print(comment_1,comment_2,comment_3,comment_4,comment_5,comment_6)
        if len(comment_1)>0:
            comment_1="<br>".join(ele for ele in comment_1)
        if len(comment_2)>0:
            comment_2="<br>".join(ele for ele in comment_2)
        if len(comment_3)>0:
            comment_3="<br>".join(ele for ele in comment_3)
        if len(comment_4)>0:
            comment_4="<br>".join(ele for ele in comment_4)   
        if len(comment_5)>0:
            comment_5="<br>".join(ele for ele in comment_5)         
        if len(comment_6)>0:
            comment_6="<br>".join(ele for ele in comment_6)               
        doc.add_comment("Comment", "<b>{0}</b><br><hr><b>{1}</b><br><hr><b>{2}</b><br><hr><b>{3}</b><br><hr><b>{4}</b><br><hr><b>{5}</b><br><hr>".format(comment_1,comment_2,comment_3,comment_4,comment_5,comment_6))
        frappe.msgprint(_("Purchase Invoice {0} created successfully").format( frappe.bold(get_link_to_form("Purchase Invoice", doc.name))))
        coal_settings = frappe.get_doc('Coal Settings')
        coal_settings.add_comment("Comment", "<b>Purchase Invoice {0} created successfully</b>".format(get_link_to_form("Purchase Invoice", doc.name)))
        # return doc.name
    except Exception as e:
            error_log=frappe.log_error(
                title="Error: create_pi_for_royalty_charges",
                message=frappe.get_traceback(),
            )
            coal_settings = frappe.get_doc('Coal Settings')
            coal_settings.add_comment("Comment", "Error in creating Purchase Invoice for Royalty Charges. To check error click - {0}".format(frappe.bold(get_link_to_form("Error Log", error_log.name))))
            frappe.db.commit()
            frappe.throw(_("Error in creating Purchase Invoice for Royalty Charges. <br> Please check logs with title create_pi_for_royalty_charges"))

def get_cargo_handling_qty_for_royalty_purchase_invoice(posting_date):
        from frappe.utils.data import get_date_str

        month_start_date=get_first_day(posting_date)
        port_start_date_with_time=get_date_str(month_start_date)+' 06:00:01'

        next_month_start_date = add_days(get_last_day(posting_date),1)
        port_next_month_date_as_end_date_with_time=get_date_str(next_month_start_date)+' 06:00:00'

        query = frappe.db.sql(
        """			select 
                sle.vessel,
                sle.item_code,		
                sum(batch_package.qty) as actual_qty				
            from
                `tabStock Ledger Entry` sle
            inner join `tabSerial and Batch Entry` batch_package
            on
                sle.serial_and_batch_bundle = batch_package.parent
            inner join `tabBatch` batch 
            on batch_package.batch_no =batch.name 
            where
                sle.docstatus < 2
                and sle.is_cancelled = 0
                and sle.has_batch_no = 1
                and actual_qty >0
                and sle.posting_datetime >='{0}'
                and sle.posting_datetime <='{1}'
            group by
                sle.vessel ,sle.item_code """.format(port_start_date_with_time,port_next_month_date_as_end_date_with_time),as_dict=1,debug=1)	
        data=[]
        vessel_done=[]
        if len(query)>0:
            for row in query:
                current_vessel=row.vessel
                current_vessel_qty=0
                current_vessel_item=[]
                if current_vessel not in vessel_done:
                    for inner_row in query:
                        if inner_row.vessel==current_vessel:
                            current_vessel_qty=current_vessel_qty+inner_row.actual_qty
                            current_vessel_comodity = frappe.db.get_value('Item', inner_row.item_code, 'custom_coal_commodity')
                            if current_vessel_comodity == None:
                                frappe.throw(_("Please set commodity in item {0}").format(inner_row.item_code))
                            else:
                                if current_vessel_comodity not in current_vessel_item:
                                    current_vessel_item.append(current_vessel_comodity)							
                    vessel_done.append(current_vessel)	
                    data.append({'vessel':current_vessel,'qty':current_vessel_qty,'item':",".join(current_vessel_item)})
        return data

def get_item_price_list_rate(vessel,royalty_invoice_item,price_list,transaction_date=None):
    vessel_doc = frappe.get_doc("Vessel",vessel)
    item_args=frappe._dict({'item_code':royalty_invoice_item,'buying_price_list':price_list,'company':vessel_doc.company,"doctype":"Purchase Invoice"})
    item_defaults=get_basic_details(item_args,item=None)
    args = frappe._dict({'price_list':price_list or 'Standard Buying','qty':1,'uom':item_defaults.get('stock_uom'),'transaction_date':transaction_date})
    price_list_rate=get_price_list_rate_for(args,royalty_invoice_item)
    return price_list_rate

def validate_vessel_is_present_in_items(self,method):
    cargo_received,handling_loss,audit_shortage = frappe.db.get_value("Coal Settings","Coal Settings",["cargo_received","handling_loss","audit_shortage"])
    stock_entry_type = self.get("stock_entry_type")
    if self.doctype=='Stock Entry' and stock_entry_type ==cargo_received or stock_entry_type ==handling_loss or stock_entry_type ==audit_shortage:
        if self.stock_entry_type==cargo_received:
            for row in self.items:
                if (not row.to_vessel) or (row.to_vessel==''):
                    frappe.throw(_("Target vessel is missing for row {0}.".format(frappe.bold(row.idx))))
        if stock_entry_type ==handling_loss or stock_entry_type ==audit_shortage:
            for row in self.items:
                if (not row.vessel) or (row.vessel==''):
                    frappe.throw(_("Source vessel is missing for row {0}.".format(frappe.bold(row.idx))))
    

    if self.doctype=='Delivery Note':
        for row in self.items:
            is_customer_provided_item = frappe.db.get_value('Item', row.item_code, 'is_customer_provided_item')
            if is_customer_provided_item==1:
                if (not row.vessel) or (row.vessel==''):
                    frappe.throw(_("Source vessel is missing for row {0}.".format(frappe.bold(row.idx))))				

    
def get_data_from_royalty_charges_report_for_vessel(posting_date,type="PI"):
    vessel_list = []
    month_start_date = get_first_day(posting_date)
    month_end_date = get_last_day(posting_date)

    query = frappe.db.sql(
        """SELECT
            DISTINCT custom_vessel
        FROM
            `tabBatch`
        WHERE
            disabled = 0
            and manufacturing_date >= '{0}'
            and manufacturing_date <= '{1}'
            and custom_vessel is not NULL
            order by manufacturing_date asc
        """.format(month_start_date,month_end_date),as_dict=1,debug=1)

    
    for row in query:
        vessel_list.append(row.custom_vessel)

    non_closed_vessel = get_all_non_closed_vessel(posting_date)
    print(non_closed_vessel,"case 2--------------non closed")
    print(vessel_list,"case 1------------ vessel from SE")

    vessel_set = set(vessel_list)

    non_closed_vessel_set = set(non_closed_vessel)
    distinct_vessel_list = list(vessel_set.union(non_closed_vessel_set))
    report_data=[]
    charged_vessel = []
    free_vessel_list = []
    print(distinct_vessel_list,"vessel")
    for vessel in distinct_vessel_list:
        vessel_data,free_vessel = get_amount_from_royalty_charges_report(vessel,month_start_date,month_end_date)
        for row in vessel_data:
            report_data.append(row)
            if row.get("vessel") not in charged_vessel:
                charged_vessel.append(row.get("vessel"))
        if len(free_vessel)>0:
            free_vessel_list.append(free_vessel[0])

    print(free_vessel_list,"freeeee")
    print(charged_vessel,"chargeddddddddd")
    if type=="PI":
        return report_data
    elif type=="Comment":
        return distinct_vessel_list,free_vessel_list,charged_vessel

def get_all_non_closed_vessel(posting_date):
    non_closed_vessels = frappe.db.get_all("Vessel",
                                        filters={"vessel_closure":0},
                                        fields=["name"],
                                        order_by="modified asc")
    non_closed_vessel_list = []
    for vessel in non_closed_vessels:
        first_line_ashore_of_vessel = getdate(frappe.db.get_value("Statement of Fact",vessel.name,"first_line_ashore"))
        date = get_last_day(getdate(posting_date))
        if first_line_ashore_of_vessel <= date:
            non_closed_vessel_list.append(vessel.name)
    return non_closed_vessel_list

def get_amount_from_royalty_charges_report(vessel,start_date,end_date):
    from kict.kict.report.royalty_storage.royalty_storage import execute
    first_line_ashore_of_vessel = frappe.db.get_value("Statement of Fact", vessel, "first_line_ashore")
    filter_for_royalty =frappe._dict({"vessel":vessel,"from_date":getdate(first_line_ashore_of_vessel),"to_date":getdate(end_date),"only_for_royalty":1})
    unique_items=get_unique_item(vessel)
    print(unique_items,"items")
    first_slot_item,first_slot_storage_charges,second_slot_item,second_slot_storage_charges=get_royalty_storage_items_and_rate(type="PI")
    col,report_data = execute(filter_for_royalty)
    get_filtered_data = []
    free_vessel = []
    slot_charges=[first_slot_storage_charges,second_slot_storage_charges]
    for item in unique_items:
        for slot_charge in slot_charges:
            qty=0
            for row in report_data:
                if row.datewise>=start_date and row.rate==slot_charge and row.bal_qty>0 and row.item_code==item.item: 
                    qty=qty+row.bal_qty
            if slot_charge==first_slot_storage_charges:
                storage_item=first_slot_item
            elif slot_charge==second_slot_storage_charges:
                storage_item=second_slot_item
            if qty>0:               
                get_filtered_data.append({"vessel":vessel,"customer_item":item.item,"qty":qty,"storage_item":storage_item})
   
    if len(get_filtered_data)==0:
        free_vessel.append(vessel)
    return get_filtered_data,free_vessel

def get_vessel_grade_details(vessel_name):
    grade_list=frappe.db.get_all('Vessel Details', filters={'parent': vessel_name},fields=['grade'])
    
    if grade_list:
        unique_grade=[]
        for grade in grade_list:
            if grade.grade not in unique_grade:
                unique_grade.append(grade.grade)
        # print(unique_grade, '--unique_grade')
        print(",".join(unique_grade), '----------')
        return ",".join(unique_grade)
        # return ', '.join( ele.grade for ele in grade_list)
    else:
        return ''
    
def validate_item_qty(self,method):
    if self.stock_entry_type == "Cargo Received":
        if len(self.items)>0:
            for row in self.items:
                if row.qty > 0:        
                    qty_from_vessel = frappe.db.get_value("Vessel Details",{"parent":self.custom_vessel,"item":row.item_code},["tonnage_mt"])
                                    
                    details = frappe.db.sql("""
                        SELECT
                            se.custom_vessel ,
                            sed.item_code ,
                            sum(sed.qty) as qty
                        FROM
                            `tabStock Entry` se
                        inner join `tabStock Entry Detail` sed on
                            se.name = sed.parent
                        where
                            se.docstatus = 1
                            and sed.item_code = '{0}'
                            and se.custom_vessel = '{1}'
                            and se.stock_entry_type = 'Cargo Received'
                        group by
                            se.custom_vessel ,
                            sed.item_code
                                            """.format(row.item_code,self.custom_vessel),debug=1,as_dict=1)
                    if len(details)>0:
                        qty_in_warehouse = details[0].qty
                        total_qty = qty_in_warehouse+row.qty
                        if qty_from_vessel and total_qty > qty_from_vessel:
                            if qty_from_vessel-qty_in_warehouse == 0:
                                frappe.throw(_("You cannot create Stock Entry for Cargo Receive because qty in warehouses cannot be greater than total tonnage"))
                            else :
                                frappe.throw(_("Row #{0}:You canot add qty greater than <b>{1}</b>".format(row.idx,qty_from_vessel-qty_in_warehouse)))
                                
                    if qty_from_vessel and qty_from_vessel<row.qty:
                        frappe.throw(_("Row #{0}: Qty cannot be greater than received qty <b>{1}</b>".format(row.idx,qty_from_vessel)))


def vessel_lay_time_calculation_sheet(vessel_name):
    # print('vessel_name',vessel_name)
    query = frappe.db.sql(
        """
        SELECT
        *
        FROM `tabStatement of Fact` as sof
        inner join `tabVessel` as vessel on
            vessel.name=sof.name
        inner join `tabVessel Details` as vessel_details on
            vessel.name = vessel_details.parent
        where
            sof.name = '{0}'
    """.format(vessel_name), as_dict=1, debug=1)
    # print(query,'query')
    return query

def check_posting_date_time_and_batch_date(self,method):
    if self.stock_entry_type == "Cargo Received":
        port_date = get_port_date(self.posting_date,self.posting_time)
        port_date_time = getdate(port_date)
        if len(self.items)>0:
            for row in self.items:
               batch_date = frappe.db.get_value("Batch",row.batch_no,"manufacturing_date")
               if port_date_time != batch_date:
                   frappe.throw(_("Stock entry port date is {0}.<br>Row #{1}: Btach {2} has port date as {3}.<br>It is not matching").format(port_date,row.idx,row.batch_no,batch_date))

def bank_address(bank_account):
    # bank_name = frappe.db.get_value("Bank Account", bank_account, "bank")
    bank_address_id = frappe.db.get_all("Dynamic Link", parent_doctype="Address",
                                        filters={"link_doctype":"Bank Account","link_name":bank_account},
                                        fields=["parent"])
    
    custom_addr = ''
    if len(bank_address_id)>0:
        doc = frappe.get_doc('Address', bank_address_id[0].parent)
        custom_addr = "{0} <br>{1}<br>{2}<br>{3} - {4}".format(doc.address_title,doc.address_line1,doc.address_line2,doc.state,doc.pincode)
        
    return custom_addr

def get_sbi_non_sbi_data(docname):
    sbi_bank_data = frappe.db.sql("""
                    SELECT
                        por.amount,
                        ba.bank_account_no,
                        ba.branch_code,
                        ba.account_name,
                        por.custom_remarks
                    FROM
                        `tabPayment Order Reference` as por
                    LEFT OUTER JOIN `tabBank Account` as ba on
                        por.bank_account = ba.name
                    WHERE ba.bank = 'State Bank of India' and por.parent = '{0}'
                """.format(docname),as_dict=True,debug=1)
    sbi_amount = 0
    if len(sbi_bank_data)>0:
        for data in sbi_bank_data:
            sbi_amount = sbi_amount + data.amount
    
    non_sbi_bank_data = frappe.db.sql("""
                SELECT
                    por.amount,
                    ba.bank_account_no,
                    ba.branch_code,
                    ba.account_name,
                    por.custom_remarks
                FROM
                    `tabPayment Order Reference` as por
                LEFT OUTER JOIN `tabBank Account` as ba on
                    por.bank_account = ba.name
                WHERE ba.bank != 'State Bank of India' and por.parent = '{0}'
            """.format(docname),as_dict=True,debug=1)
    non_sbi_amount = 0
    if len(non_sbi_bank_data)>0:
        for row in non_sbi_bank_data:
            non_sbi_amount = non_sbi_amount + row.amount
    
    total_amount = sbi_amount + non_sbi_amount
    print(sbi_amount, non_sbi_amount, total_amount, "sbi_amount, non_sbi_amount, total_amount")
    print(sbi_bank_data, non_sbi_bank_data)
    return sbi_amount, non_sbi_amount, total_amount, sbi_bank_data, non_sbi_bank_data

@frappe.whitelist()
def delete_file(file_name):
    updated_file_name = file_name.replace("/","-")
    print("+"*100)
    public_file_path = frappe.get_site_path("public", "files")
    file_path = os.path.join(public_file_path, updated_file_name)
    if os.path.exists(file_path):
        os.remove(file_path)
        return True
    else:
        return False

@frappe.whitelist()
def get_sbi_account_details(docname,file_name):
    updated_file_name = file_name.replace("/","-")
    file_header = [
        
            "IFSC CODE",
            "AMOUNT",
            "BENEFICIARY AC NO",
            "BENFICIARY NAME",
            "BENEFICIARY ADDRESS"
        
    ]
    print(file_header,"file_header")

    po_data = frappe.db.sql("""
                    SELECT
                        ba.branch_code,
                        por.amount,
                        ba.bank_account_no,
                        s.supplier_name,
                        ad.city
                    FROM
                        `tabPayment Order Reference` as por
                    INNER JOIN `tabSupplier` as s on 
                        s.name = por.supplier
                    LEFT OUTER JOIN `tabBank Account` as ba on
                        por.bank_account = ba.name
                    LEFT OUTER JOIN `tabDynamic Link` as dl on
                        ba.name = dl.link_name
                    LEFT OUTER JOIN `tabAddress` as ad on
                        ad.name = dl.parent
                    WHERE ba.bank = 'State Bank of India' and por.parent = '{0}'
                """.format(docname),as_dict = 1)
    print(po_data,"po_data",type(po_data))

    total_amount = 0
    if len(po_data)>0:
        for row in po_data:
            total_amount = total_amount + row.amount

    file_footer = [
        "Total Amount",
        total_amount,
        "",
        "",
        ""
    ]
    # xlsx file creation

    public_file_path = frappe.get_site_path("public", "files")
    workbook = openpyxl.Workbook(write_only=True)
    # file_name=f"SBI-{docname}.xlsx"
    file_url=os.path.join(public_file_path,updated_file_name)
    sheet = workbook.create_sheet("SBI", 0)
    sheet.append(file_header)
    for ele in po_data:
        sheet.append([ele.branch_code,ele.amount,ele.bank_account_no,ele.supplier_name,ele.city])
    sheet.append(file_footer)
    workbook.save(file_url)

    # excel fromatting

    workBook = openpyxl.load_workbook(file_url)
    workSheet = workBook.active

    for i in range(1, workSheet.max_column + 1):
        workSheet.cell(1, i).font = Font(bold=True, size=12, name="Calibri")
        workSheet.cell(workSheet.max_row, i).font = Font(bold=True, size=10, name="Calibri")
        workSheet.row_dimensions[1].height = 20
        # workSheet.column_dimensions['A'].width = 20
        # workSheet.column_dimensions['B'].width = 20
        # workSheet.column_dimensions['C'].width = 35
        # workSheet.column_dimensions['D'].width = 35
        # workSheet.column_dimensions['E'].width = 30
    
    for column_cells in workSheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (chr(64+(column_cells[0].column)))
        if new_column_length > 0:
            workSheet.column_dimensions[new_column_letter].width = new_column_length+5

    border_thin = Side(style='thin')
    for i in range (1, workSheet.max_row + 1):
        for j in range(1, workSheet.max_column + 1):
            workSheet.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
            workSheet.cell(i, j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

    workBook.save(file_url)
    return frappe.utils.get_url()+"/files/"+updated_file_name

@frappe.whitelist()
def get_non_sbi_account_details(docname,file_name):
    updated_file_name = file_name.replace("/","-")

    file_header = [
        "IFSC CODE",
        "AMOUNT",
        "BENEFICIARY AC NO",
        "BENFICIARY NAME",
        "BENEFICIARY ADDRESS"
    ]

    po_data = frappe.db.sql("""
                    SELECT
                        ba.branch_code,
                        por.amount,
                        ba.bank_account_no,
                        s.supplier_name,
                        ad.city
                    FROM
                        `tabPayment Order Reference` as por
                    INNER JOIN `tabSupplier` as s on 
                        s.name = por.supplier 
                    LEFT OUTER JOIN `tabBank Account` as ba on
                        por.bank_account = ba.name
                    LEFT OUTER JOIN `tabDynamic Link` as dl on
                        ba.name = dl.link_name
                    LEFT OUTER JOIN `tabAddress` as ad on
                        ad.name = dl.parent
                    WHERE ba.bank != 'State Bank of India' and por.parent = '{0}'
                """.format(docname),as_dict = 1)
    print(po_data,"po_data",type(po_data))
    
    total_amount = 0
    if len(po_data)>0:
        for row in po_data:
            total_amount = total_amount + row.amount
    
    file_footer = [
        "Total Amount",
        total_amount,
        "",
        "",
        ""
    ]

    # xlsx file creation

    public_file_path = frappe.get_site_path("public", "files")
    workbook = openpyxl.Workbook(write_only=True)
    # file_name=f"Non-SBI-{docname}.xlsx"
    file_url=os.path.join(public_file_path,updated_file_name)
    sheet = workbook.create_sheet("Non-SBI", 0)
    sheet.append(file_header)
    for ele in po_data:
        sheet.append([ele.branch_code,ele.amount,ele.bank_account_no,ele.supplier_name,ele.city])
    sheet.append(file_footer)
    workbook.save(file_url)

    workBook = openpyxl.load_workbook(file_url)
    workSheet = workBook.active

    for i in range(1, workSheet.max_column + 1):
        workSheet.cell(1, i).font = Font(bold=True, size=12, name="Calibri")
        workSheet.cell(workSheet.max_row, i).font = Font(bold=True, size=10, name="Calibri")
        workSheet.row_dimensions[1].height = 20
        # workSheet.column_dimensions['A'].width = 20
        # workSheet.column_dimensions['B'].width = 20
        # workSheet.column_dimensions['C'].width = 40
        # workSheet.column_dimensions['D'].width = 30
        # workSheet.column_dimensions['E'].width = 30
    
    for column_cells in workSheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (chr(64+(column_cells[0].column)))
        if new_column_length > 0:
            workSheet.column_dimensions[new_column_letter].width = new_column_length+5

    border_thin = Side(style='thin')
    for i in range (1, workSheet.max_row + 1):
        for j in range(1, workSheet.max_column + 1):
            workSheet.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
            workSheet.cell(i, j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

    workBook.save(file_url)
    return frappe.utils.get_url()+"/files/"+updated_file_name

@frappe.whitelist()
def get_statement_with_remarks_data(docname,file_name):
    updated_file_name = file_name.replace("/","-")
    file_header = [
        "Sl No",
        "Name of the Party",
        "Amount Rs.",
        "Bank Type",
        "Description",
        "On Account of"
    ]
    print(file_header,"file_header")

    po_data = frappe.db.sql("""
                        SELECT
                            s.supplier_name,
                            por.amount,
                            ba.bank,
                            por.custom_remarks
                        FROM
                            `tabPayment Order Reference` as por
                        inner join `tabSupplier` as s 
                            on por.supplier = s.name
                        LEFT OUTER JOIN `tabBank Account` as ba 
                            on por.bank_account = ba.name 
                        WHERE por.parent = '{0}'
                """.format(docname),as_dict = 1)
    
    print(po_data,"po_data",type(po_data))
    updated_data = []
    idx = 1
    total_amount = 0
    if len(po_data)>0:
        for row in po_data:
            row_idx = list(row.keys()).index('supplier_name')
            items = list(row.items())
            items.insert(row_idx, ('sr_no',idx))
            row = dict(items)
            total_amount = total_amount + row.get("amount")
            if row.get("bank") == "State Bank of India":
                row.update({"bank": "SBI"})
            else :
                row.update({"bank": "Non-SBI"})
            row["on_account_of"] = "KICTPPL"
            idx = idx + 1
            print(row,"---")
            updated_data.append(row)
    print(updated_data,"+++===+++")

    file_footer = [
            "",
            "Payout(A)",
            total_amount,
            "",
            "",
            ""
        ]

    # xlsx file creation

    public_file_path = frappe.get_site_path("public", "files")
    workbook = openpyxl.Workbook(write_only=True)
    # file_name=f"Statement-with-remarks-{docname}.xlsx"
    file_url=os.path.join(public_file_path,updated_file_name)
    sheet = workbook.create_sheet("Statement With Remarks", 0)
    sheet.append(file_header)
    for ele in updated_data:
        sheet.append([ele.get("sr_no"),ele.get("supplier_name"),ele.get("amount"),ele.get("bank"),ele.get("custom_remarks"),ele.get("on_account_of")])
    sheet.append(file_footer)
    workbook.save(file_url)

    # file formatting
    workBook = openpyxl.load_workbook(file_url)
    workSheet = workBook.active

    for i in range(1, workSheet.max_column + 1):
        workSheet.cell(1, i).font = Font(bold=True, size=12, name="Calibri")
        workSheet.cell(workSheet.max_row, i).font = Font(bold=True, size=10, name="Calibri")
        workSheet.row_dimensions[1].height = 20
        # workSheet.column_dimensions['A'].width = 15
        # workSheet.column_dimensions['B'].width = 30
        # workSheet.column_dimensions['C'].width = 20
        # workSheet.column_dimensions['D'].width = 20
        # workSheet.column_dimensions['E'].width = 30
        # workSheet.column_dimensions['F'].width = 20
    
    for column_cells in workSheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (chr(64+(column_cells[0].column)))
        if new_column_length > 0:
            workSheet.column_dimensions[new_column_letter].width = new_column_length+5

    border_thin = Side(style='thin')
    for i in range (1, workSheet.max_row + 1):
        for j in range(1, workSheet.max_column + 1):
            workSheet.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
            workSheet.cell(i, j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

    workBook.save(file_url)

    return frappe.utils.get_url()+"/files/"+updated_file_name

@frappe.whitelist()
def get_purchase_invoice_data(docname,file_name):
    updated_file_name = file_name.replace("/","-")
    file_header = [
        ["","","","","KALINGA INTERNATIONAL COAL TERMINAL PARADIP PVT LTD (KICTPPL)"],
        ["","","","","List of Local expenses 2024"],
        [
            "Sr No",
            "Date of Invoices",
            "Invoice No.",
            "Date of Certificatation(MSME vendor)",
            "Vendor",
            "Description",
            "Basic Taxable Amount",
            "Non-taxable Amount",
            "GST",
            "TDS",
            "Retention Labour Cess",
            "Retention Money",
            "Adjustment",
            "Net Payble"
    ]
    ]
    print(file_header,"file header -------------")

    pi_data = frappe.db.sql("""
                    SELECT
                        pi.posting_date,
                        pi.bill_date,
                        pi.name,
                        s.supplier_name,
                        por.custom_remarks,
                        pi.itc_integrated_tax,
                        pi.itc_central_tax,
                        pi.itc_state_tax,
                        pr.custom_adjustment
                    FROM
                        `tabPayment Order Reference` por
                    inner join `tabPurchase Invoice` pi on
                        pi.name = por.reference_name
                    inner join `tabSupplier` s on
                        s.name = por.supplier
                    inner join `tabPayment Request` pr on
                        pr.name = por.payment_request
                    WHERE por.parent = '{0}'
            """.format(docname),as_dict = 1, debug=1)
    print(pi_data,"pi data ==========")

    new_row = {}
    updated_data = []
    idx = 1

    account_head_for_retention_labour_cess = frappe.db.get_single_value("Coal Settings","account_head_for_retention_labour_cess")
    account_head_for_retention_money = frappe.db.get_single_value("Coal Settings","account_head_for_retention_money")

    basic_tax_total = 0
    non_tax_total = 0
    gst_total = 0
    tds_total = 0
    net_payable = 0
    total_retention_labour_cess_amount = 0
    total_retention_money = 0
    total_adjustment = 0

    for row in pi_data:
        print(row.name)
        taxable_amount = 0
        non_taxable_amount = 0
        tds_amount = 0
        total_gst = 0
        pi_doc = frappe.get_doc("Purchase Invoice",row.name)
        if len(pi_doc.items)>0:
            for pi_child in pi_doc.items:
                if pi_child.gst_treatment == "Taxable":
                    taxable_amount = taxable_amount + pi_child.amount
                elif pi_child.gst_treatment == "Nil-Rated":
                    non_taxable_amount = non_taxable_amount + pi_child.amount

        total_gst = (pi_doc.itc_integrated_tax or 0) + (pi_doc.itc_central_tax or 0) + (pi_doc.itc_state_tax or 0)
        retention_labour_cess_amount = 0
        retention_money = 0

        if len(pi_doc.taxes)>0:
            for tax_row in pi_doc.taxes:
                parent_of_account_head = frappe.db.get_value("Account",tax_row.account_head,"parent_account")
                if parent_of_account_head and parent_of_account_head == account_head_for_retention_labour_cess:
                    retention_labour_cess_amount = retention_labour_cess_amount + tax_row.tax_amount
                if parent_of_account_head and parent_of_account_head == account_head_for_retention_money:
                    retention_money = retention_money + tax_row.tax_amount
                if tax_row.is_tax_withholding_account == 1:
                    tds_amount = tds_amount + tax_row.tax_amount

        new_row = {}
        new_row["sr_no"] = idx
        new_row["date_of_invoice"] = row.bill_date
        new_row["invoice_no"] = row.name
        new_row["msme_date"] = row.posting_date
        new_row["supplier"] = row.supplier_name
        new_row["description"] = row.custom_remarks
        new_row["basic_taxable_amount"] = taxable_amount if taxable_amount > 0 else "-"
        new_row["non_taxable_amount"] = non_taxable_amount if non_taxable_amount > 0 else "-"
        new_row["gst_amount"] = total_gst if total_gst > 0 else "-"
        new_row["tds_amount"] = tds_amount if tds_amount > 0 else "-"
        new_row["retention_labour_cess_amount"] = retention_labour_cess_amount or "-"
        new_row["retention_money"] = retention_money or "-"
        new_row["adjustment"] = row.custom_adjustment if row.custom_adjustment else ""
        new_row["net_payable_amount"] = pi_doc.grand_total

        basic_tax_total = basic_tax_total + (taxable_amount or 0)
        non_tax_total = non_tax_total + (non_taxable_amount or 0)
        gst_total = gst_total + (total_gst or 0)
        tds_total = tds_total + (tds_amount or 0)
        net_payable = net_payable + pi_doc.grand_total
        total_retention_labour_cess_amount = total_retention_labour_cess_amount + (retention_labour_cess_amount or 0)
        total_retention_money = total_retention_money + (retention_money or 0)
        total_adjustment = total_adjustment + (row.custom_adjustment or 0)
    
        updated_data.append(new_row)
        idx = idx + 1
    
    po_data = frappe.db.sql("""
                    SELECT
                        po.transaction_date,
                        po.name,
                        s.supplier_name,
                        por.custom_remarks,
                        por.amount,
                        pr.custom_basic_taxable_amount as taxable_amount,
                        pr.custom_nontaxable_amount as non_taxable_amount,
                        pr.custom_gst_amount as total_gst,
                        pr.custom_tds as tds_amount,
                        pr.custom_retention_labour_cess as retention_labour_cess_amount,
                        pr.custom_retention_money as retention_money,
                        pr.custom_adjustment                       
                    FROM
                        `tabPayment Order Reference` por
                    inner join `tabPurchase Order` po on
                        po.name = por.reference_name
                    inner join `tabSupplier` s on
                        s.name = por.supplier
                    inner join `tabPayment Request` pr on
                        pr.name = por.payment_request
                    WHERE por.parent = '{0}'
            """.format(docname),as_dict = 1, debug=1)
    print(po_data,"po data ==========")
    if len(po_data)>0:
        for po in po_data:
            new_po_row = {}
            new_po_row["sr_no"] = idx
            new_po_row["date_of_invoice"] = po.transaction_date
            new_po_row["invoice_no"] = po.name
            new_po_row["msme_date"] = "-"
            new_po_row["supplier"] = po.supplier_name
            new_po_row["description"] = po.custom_remarks
            new_po_row["basic_taxable_amount"] = po.taxable_amount if po.taxable_amount > 0 else "-"
            new_po_row["non_taxable_amount"] = po.non_taxable_amount if po.non_taxable_amount > 0 else "-"
            new_po_row["gst_amount"] = po.total_gst if po.total_gst > 0 else "-"
            new_po_row["tds_amount"] = po.tds_amount if po.tds_amount > 0 else "-"
            new_po_row["retention_labour_cess_amount"] = po.retention_labour_cess_amount or "-"
            new_po_row["retention_money"] = po.retention_money or "-"
            new_po_row["adjustment"] = po.custom_adjustment if po.custom_adjustment else ""
            new_po_row["net_payable_amount"] = po.amount

            basic_tax_total = basic_tax_total + (po.taxable_amount or 0)
            non_tax_total = non_tax_total + (po.non_taxable_amount or 0)
            gst_total = gst_total + (po.total_gst or 0)
            tds_total = tds_total + (po.tds_amount or 0)
            net_payable = net_payable + po.amount
            total_retention_labour_cess_amount = total_retention_labour_cess_amount + (po.retention_labour_cess_amount or 0)
            total_retention_money = total_retention_money + (po.retention_money or 0)
            total_adjustment = total_adjustment + (po.custom_adjustment or 0)
        
            updated_data.append(new_po_row)
            idx = idx + 1

    # total_row = {}
    # total_row["description"] = "Total"
    # total_row["basic_taxable_amount"] = basic_tax_total
    # total_row["non_taxable_amount"] = non_tax_total
    # total_row["gst_amount"] = gst_total
    # total_row["tds_amount"] = tds_total
    # total_row["net_payable_amount"] = net_payable

    # updated_data.append(total_row)

    file_footer = [
            "",
            "",
            "",
            "",
            "",
            "Total",
            basic_tax_total,
            non_tax_total,
            gst_total,
            tds_total,
            total_retention_labour_cess_amount,
            total_retention_money,
            total_adjustment,
            net_payable
        ]
       
    print(updated_data,"))))))))))))")

    # xlsx file creation

    public_file_path = frappe.get_site_path("public", "files")
    workbook = openpyxl.Workbook(write_only=True)
    # file_name=f"Purchase-Invoice-{docname}.xlsx"
    file_url=os.path.join(public_file_path,updated_file_name)
    sheet = workbook.create_sheet("Invoices", 0)
    for header in file_header:
        sheet.append(header)
    for ele in updated_data:
        sheet.append([ele.get("sr_no"),ele.get("date_of_invoice"),ele.get("invoice_no"),ele.get("msme_date"),ele.get("supplier"),ele.get("description"),ele.get("basic_taxable_amount"),
                      ele.get("non_taxable_amount"),ele.get("gst_amount"),ele.get("tds_amount"),ele.get("retention_labour_cess_amount"),ele.get("retention_money"),ele.get("adjustment"),ele.get("net_payable_amount")])
    sheet.append(file_footer)
    workbook.save(file_url)

    # file Formatting
    workBook = openpyxl.load_workbook(file_url)
    workSheet = workBook.active

    for i in range(1, workSheet.max_column + 1):
        workSheet.cell(1, i).font = Font(bold=True, size=12, name="Calibri", underline='single')
        workSheet.row_dimensions[1].height = 25
        workSheet.cell(2, i).font = Font(bold=True, size=10, name="Calibri", underline='single')
        workSheet.row_dimensions[2].height = 20
        workSheet.cell(3, i).font = Font(bold=True, size=10, name="Calibri")

        # workSheet.column_dimensions['A'].width = 15
        # workSheet.column_dimensions['B'].width = 30
        # workSheet.column_dimensions['C'].width = 20
        # workSheet.column_dimensions['D'].width = 20
        # workSheet.column_dimensions['E'].width = 30
        # workSheet.column_dimensions['F'].width = 20
        # workSheet.column_dimensions['G'].width = 25
        # workSheet.column_dimensions['H'].width = 25
        # workSheet.column_dimensions['I'].width = 20
        # workSheet.column_dimensions['J'].width = 20
        # workSheet.column_dimensions['K'].width = 30
        # workSheet.column_dimensions['L'].width = 20
        # workSheet.column_dimensions['M'].width = 20
        # workSheet.column_dimensions['N'].width = 20
    
    for column_cells in workSheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (chr(64+(column_cells[0].column)))
        if new_column_length > 0:
            workSheet.column_dimensions[new_column_letter].width = new_column_length+5    
    
    workSheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=workSheet.max_column)
    row1 = workSheet.cell(row = 1, column = 1) 
    row1.value = "KALINGA INTERNATIONAL COAL TERMINAL PARADIP PVT LTD (KICTPPL)"

    workSheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=workSheet.max_column)
    row2 = workSheet.cell(row = 2, column = 1)
    row2.value="List of Local Expenses"

    border_thin = Side(style='thin')
    print(workSheet.max_column, "---workSheet.max_column---", workSheet.max_row, "---workSheet.max_row----")
    for i in range (1, workSheet.max_row + 1):
        for j in range(1, workSheet.max_column + 1):
            workSheet.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
            workSheet.cell(i, j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

            if (workSheet.max_row) == i:
                    workSheet.cell(i, j).font = Font(bold=True, size=11, name="Calibri")

    workBook.save(file_url)

    return frappe.utils.get_url()+"/files/"+updated_file_name